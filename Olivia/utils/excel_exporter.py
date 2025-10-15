import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
from zoneinfo import ZoneInfo


def export_to_excel(queryset, headers, row_data_func, file_prefix="export"):
    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Define styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="000000")
    body_font = Font(name="Calibri", size=11)
    alignment = Alignment(vertical="center", wrap_text=False)

    # Border style
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Write header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # Write data rows
    for row_num, obj in enumerate(queryset, 2):
        row = row_data_func(obj)
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = body_font
            cell.alignment = alignment
            cell.border = thin_border

    # Auto-fit column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            (len(str(ws.cell(row=row, column=col_num).value or "")) for row in range(1, ws.max_row + 1)),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Save workbook to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Saudi Arabia time + add 3 hours
    saudi_arabia_tz = ZoneInfo("Asia/Riyadh")
    now_saudi_arabia = datetime.datetime.now(saudi_arabia_tz) + datetime.timedelta(hours=3)

    # Format for Windows-safe filename (replace ":" with "-")
    timestamp = now_saudi_arabia.strftime("%Y-%m-%d %H-%M-%S")

    filename = f"{file_prefix}_{timestamp}_KSA.xlsx"

    # Response
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
